import time
import json
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

class CloudAuditTrailTracker:
    def __init__(self):
        self.log = []

    def log_event(self, user, action, status="Success", system_event=False):
        event = {
            "timestamp": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "user": user if not system_event else "System",
            "action": action,
            "status": status
        }
        self.log.append(event)
        print(f"Event logged: {event}")

    def generate_report(self, user=None):
        if user:
            print(f"Generating report for user: {user}")
            user_log = [event for event in self.log if event["user"] == user]
            return user_log
        else:
            print("Generating full audit report")
            return self.log

    def detect_suspicious_activity(self, user):
        actions = [event for event in self.log if event["user"] == user and event["status"] == "Failed"]
        if len(actions) > 3:
            print(f"Suspicious activity detected for user: {user}")
        else:
            print(f"No suspicious activity for user: {user}")

    def save_to_excel(self, data, filename):
        wb = openpyxl.Workbook()
        ws = wb.active

  
        header_row = ["Timestamp", "User", "Action", "Status"]
        for col, value in enumerate(header_row, start=1):
            ws.cell(row=1, column=col).value = value

        
        for row, event in enumerate(data, start=2):
            ws.cell(row=row, column=1).value = event["timestamp"]
            ws.cell(row=row, column=2).value = event["user"]
            ws.cell(row=row, column=3).value = event["action"]
            ws.cell(row=row, column=4).value = event["status"]

       
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        wb.save(filename)


tracker = CloudAuditTrailTracker()

while True:
    user = input("Enter your username: ")
    action = input("Enter your action (Login, Logout, View Account Details, etc.): ")
    status = input("Enter the status of your action (Success, Failed, etc.): ")

    tracker.log_event(user=user, action=action, status=status)

    print("\nDo you want to continue? (yes/no): ")
    choice = input().lower()
    if choice != "yes":
        break


tracker.detect_suspicious_activity(user="bob")


report = tracker.generate_report()
tracker.save_to_excel(report, "full_report.xlsx")
print("Full report saved to full_report.xlsx")

user_report = tracker.generate_report(user="alice")
tracker.save_to_excel(user_report, "alice_report.xlsx")
print("User report saved to alice_report.xlsx")

